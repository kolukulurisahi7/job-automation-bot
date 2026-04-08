from nvoid_scraper import scrape_jobs
from job_parser import parse_job
from gmail_service import get_gmail_service, create_draft, get_gmail_profile
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import time
import re
import html
import config

TRACKER_FILE = config.TRACKER_FILE
MAX_EMAILS_PER_RUN = config.MAX_EMAILS_PER_RUN
DELAY_BETWEEN_EMAILS = config.DELAY_BETWEEN_EMAILS
EMPLOYER_EMAIL = config.EMPLOYER_EMAIL
KEYWORDS = config.KEYWORDS


def extract_job_id(url):
    match = re.search(r"id=(\d+)", url or "")
    return match.group(1) if match else None


def normalize_text(text: str) -> str:
    return re.sub(r"[^a-z0-9 ]+", " ", (text or "").lower()).strip()


def is_job_already_processed(job_id, title=None, description=None):
    """Check if job was already processed in previous runs."""
    if not os.path.exists(TRACKER_FILE):
        return False
    
    title_signature = normalize_text(title or "")
    desc_signature = normalize_text(description or "")
    try:
        wb = load_workbook(TRACKER_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            existing_job_id = row[0]
            existing_title = row[1] or ""
            existing_title_signature = normalize_text(existing_title)
            existing_desc_signature = ""
            if len(row) > 7 and row[7]:
                existing_desc_signature = str(row[7])

            if job_id and existing_job_id == job_id:
                return True
            if title_signature and existing_title_signature == title_signature:
                return True
            if desc_signature and existing_desc_signature == desc_signature:
                return True
        return False
    except Exception as e:
        print(f"Error checking job history: {e}")
        return False


def is_title_relevant(title: str) -> bool:
    title = (title or "").lower()
    
    has_keyword = any(word in title for word in KEYWORDS)
    has_excluded = any(word in title for word in config.EXCLUDE_KEYWORDS)
    
    return has_keyword and not has_excluded


def is_relevant_job(title: str, description: str = "") -> bool:
    title = (title or "").lower()
    description = (description or "").lower()
    
    # Check if it matches keywords
    has_keyword = any(word in title for word in KEYWORDS)
    
    # Check if it has excluded keywords in title or description
    has_excluded = any(word in title for word in config.EXCLUDE_KEYWORDS) or \
                   any(word in description for word in config.EXCLUDE_KEYWORDS)
    
    return has_keyword and not has_excluded


def pick_best_email(emails):
    """Pick the best recruiter email, avoiding generic domains."""
    if not emails:
        return None
    
    # First, try to find a non-excluded domain email
    for email in emails:
        email_lower = email.lower()
        if not any(excluded in email_lower for excluded in config.EXCLUDE_DOMAINS):
            return email
    
    # If all are excluded domains, return None
    return None


def build_email_body(title, link, recruiter_email, description):
    escaped_title = html.escape(title or "")
    escaped_link = html.escape(link or "")
    escaped_email = html.escape(recruiter_email or "")
    escaped_desc = html.escape(description or "")

    return f"""<html>
  <body style='font-family:Arial, sans-serif; font-size:14px; color:#111;'>
    <p>Hello,</p>
    <p>Hope you are doing well.</p>
    <p>This is Sahi. I've gone through the JD and it perfectly aligns with my skills. I'm a good fit for the position. I am attaching my resume and would love an opportunity to discuss how my experience matches this role.</p>
    <hr>
    <h3 style='margin-bottom:4px;'>Job Details</h3>
    <p style='margin-top:0;'>
      <strong>Title:</strong> {escaped_title}<br>
      <strong>Recruiter Email:</strong> {escaped_email}<br>
      <strong>Link:</strong> <a href="{escaped_link}">{escaped_link}</a>
    </p>
    <h4 style='margin-bottom:4px;'>Job Description</h4>
    <pre style='font-family:inherit; white-space:pre-wrap; word-break:break-word; background:#f9f9f9; padding:12px; border:1px solid #ddd; border-radius:4px;'>
{escaped_desc}
    </pre>
    <hr>
    <p>Regards,<br>Sahi</p>
  </body>
</html>"""


def ensure_tracker_file():
    if not os.path.exists(TRACKER_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["job_id", "title", "link", "emails", "phones", "status", "description_signature", "timestamp"])
        wb.save(TRACKER_FILE)


def save_applied_job(job_id, title, link, emails, phones, status, description_signature=""):
    ensure_tracker_file()

    wb = load_workbook(TRACKER_FILE)
    ws = wb.active

    ws.append([
        job_id,
        title,
        link,
        ",".join(emails),
        ",".join(phones),
        status,
        description_signature,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])

    wb.save(TRACKER_FILE)


def main():
    try:
        driver, jobs = scrape_jobs(config.SEARCH_QUERIES[0])
    except Exception as e:
        print(f"Error scraping jobs: {e}")
        return

    try:
        service = get_gmail_service()
        profile = get_gmail_profile(service)
        print(f"Authenticated Gmail: {profile.get('emailAddress')}")
        drafted_count = 0
        skipped_duplicates = 0

        print(f"\nTotal jobs found: {len(jobs)}\n")

        seen_signatures = set()
        seen_description_signatures = set()
        for i, job in enumerate(jobs, 1):
            if drafted_count >= MAX_EMAILS_PER_RUN:
                break

            title = job.get("title", "")
            link = job.get("link", "")
            job_id = extract_job_id(link)
            signature = normalize_text(title)

            print(f"\nProcessing job {i}...")
            print("Title:", title)

            # Check for duplicates by job ID or normalized title signature
            if is_job_already_processed(job_id, title):
                print("Skipped: duplicate (already processed)")
                skipped_duplicates += 1
                continue

            if not is_title_relevant(title):
                print("Skipped: not relevant by title")
                continue

            # Also avoid duplicate titles within the current run
            if signature and signature in seen_signatures:
                print("Skipped: duplicate title in current run")
                skipped_duplicates += 1
                continue

            seen_signatures.add(signature)

            try:
                parsed = parse_job(driver, job)
            except Exception as e:
                print(f"Error parsing job: {e}")
                continue

            emails = parsed.get("emails", [])
            phones = parsed.get("phones", [])
            desc = parsed.get("description", "")
            desc_signature = normalize_text(desc)

            if desc_signature and desc_signature in seen_description_signatures:
                print("Skipped: duplicate job description in current run")
                skipped_duplicates += 1
                continue

            if is_job_already_processed(job_id, title, desc):
                print("Skipped: duplicate (already processed by description)")
                skipped_duplicates += 1
                continue

            seen_description_signatures.add(desc_signature)

            print("Emails:", emails)

            if not emails:
                print("Skipped: no email")
                continue

            best_email = pick_best_email(emails)

            if not best_email:
                print("Skipped: no valid recruiter email")
                continue

            subject = title[:80]

            body = f"""Hello ,

Hope you are doing well

This is Sahi. I've gone through the JD . and it perfectly aligns with my skills .I'm a good fit for the  position. I am attaching my resume. I'm keen on discussing this opportunity further and showcasing how my skills and experience align with the requirements of the role.

{desc[:3000]}

"""

            body = build_email_body(title, link, best_email, desc)
            try:
                draft = create_draft(service, best_email, subject, body, cc=EMPLOYER_EMAIL, html=True)
                print(f"Draft created for {best_email} (id={draft.get('id')})")

                save_applied_job(job_id, title, link, [best_email], phones, "draft_created", desc_signature)
                drafted_count += 1

                time.sleep(DELAY_BETWEEN_EMAILS)

            except Exception as e:
                print(f"Draft creation failed: {e}")
                save_applied_job(job_id, title, link, emails, phones, "draft_failed")

        print(f"\nDone. Drafts created: {drafted_count}")
        if skipped_duplicates > 0:
            print(f"Duplicates skipped: {skipped_duplicates}")

    except Exception as e:
        print(f"Unexpected error in main: {e}")
    finally:
        try:
            driver.quit()
        except:
            pass


if __name__ == "__main__":
    main()