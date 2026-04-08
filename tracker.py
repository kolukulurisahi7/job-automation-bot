import os
from openpyxl import Workbook, load_workbook

FILE = "applied_jobs.xlsx"


def load_applied():
    if not os.path.exists(FILE):
        return set()

    wb = load_workbook(FILE)
    ws = wb.active

    links = set()

    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0]:
            links.add(row[0])

    return links


def save_job(link):
    if not os.path.exists(FILE):
        wb = Workbook()
        ws = wb.active
    else:
        wb = load_workbook(FILE)
        ws = wb.active

    ws.append([link])
    wb.save(FILE)